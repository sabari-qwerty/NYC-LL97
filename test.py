from openpyxl import load_workbook
from webdriver_manager.chrome import ChromeDriverManager
from selenium import webdriver
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.common.by import By


from requests import get, post


from pathlib import Path
from time import sleep
from base64 import b64decode
from bs4 import BeautifulSoup
from openpyxl import load_workbook, Workbook, drawing
from PIL import Image
import io
from execjs import compile


ld = load_workbook('data.xlsx').active

wb = Workbook()
wa = wb.active


def getDobnow(address):
    dobnow_url = "https://a810-dobnow.nyc.gov:443/Publish/WrapperPP/PublicPortal.svc/getPublicPortalPropertyDetailsGet/" + address
    dobnow_cookies = {"_gid": "GA1.2.689702731.1712843572", "RT": "\"z=1&dm=a810-dobnow.nyc.gov&si=e577c804-bfa4-4a11-b664-c04d7b66cc3b&ss=luxqrlny&sl=0&tt=0\"", "_abck": "7B9BD723430B39615EB182E1B879774C~0~YAAQh7csMWKnO5qOAQAA98bd2wsB7Mv+orUtPL+VrMepx0YFXGZNTdn178z/9vkC4ax8LVHMX/c6hIdTzjCrcAdTA2PKak0+juw9zjvSj+vqWHxy/skYJ77V5ldfi2gMBBiwJJwgNIdqOy51hDzK02XZJvaVHrgLo825kQ/WFsw15uYN5OInYAlMYZUTNEQQPVGGzchAn83W8AWnNjCKgR9RbGx0TFkQH/fGxTtp2JJXFII+0PBJCwuqR1lm64/aN/pfShqGFEndCdkiFtTXTguMRzBbObeT1wvSZyQx8r1DPB7J8n0FJObHbvlDivl4WlwuH86RTncfvOvExhoiIHBNsMciQZIGglkhisczFidgiJ1hspigWmXQdP2j+Cigo3GTY7QzXCApKHmcX/VXVYf9Or0I~-1~-1~-1", "bm_sz": "75526566572F8FF567D15EB783B3E3B6~YAAQh7csMWWnO5qOAQAA98bd2xc4eAtLYKEtiXLrJ33xktSNiZVngDELNqt2TtWNC5FLwYoJTgrHATX9j/D8c3ZS/8kzOgCSB6pPMrw2ERmmXzZ1h5PqkRsUoaB0TQ1jNmPrfIsadSs9q6kkfs0GcFyLBJFSAd4L8+pLZKeI8ZA/NieKLdR8DMGJzWF6urW5tydwFbmzNfWqTr+GvXkXsq7zyoyK19AgS7MJXyJeStawDiEPRi0cfzQ13mIhLK/Iy3zfi648JeecDZmj8PSjZ0k9f+WRt+cNfBgyeWxYP5QLWvY22sbSg73CD8CosgLa2TAMN8EsAdOFU0dVZz8XNxG8Sb3vC6O0a1GRe482mKGUVVVIIMfec0wL41JtoTsW5782TInyx8U=~3354949~4408372", "PUNKDDUF": "02ab9cfba6-09e1-43R2b4YO3D2KCbSNHCsQKIlCIHJrRbs-zcnFSIvr7oXtxX0CcG5JUfnnpVmyNv1h3Za28",
                      "bm_mi": "E903198E825FBF95CC92E2E311D083C2~YAAQh7csMW6nO5qOAQAAAdLd2xdxOvzkMWLagcXkytlk5QixShzaMncv8zECPUex7QE3+vr6G0F337d3ADRf3Bwx/tP7wEy2c8TC8Vwj8Tv6uHm5ZZcTI2PVrbHGQp84MQ3ZZ7bubifCJ8HszKf9irariI0E8o3DxIgFWpkl3SSiMxohIx4m/MHelpTAV/udgYJAGYSpu11FQgaM3xJ2aZBL7ynWIhqcUNalpC/Mm/fqeB9gKwBIPczV5Hy0xBcA4Pn3OhRvsUEM23Z6/JYYWRStiRpkZThEwgfV5V7P2fZXre2jywSrQpBtvUgVMBw4Oipip56gLjz44cai3KQ6rA==~1", "ASP.NET_SessionId": "aurylxnfhi1u4f102241bqfr", "_ga_863DM8YSJL": "GS1.1.1713085731.9.0.1713085731.0.0.0", "_ga": "GA1.2.1937995955.1712671894", "_gat_gtag_UA_128025137_1": "1", "ak_bmsc": "74E1441C2A1B8732DDBE89DF4BD640D7~000000000000000000000000000000~YAAQh7csMYGnO5qOAQAA19Pd2xd7jT7X9Ocw8RbIsIfgTWkRQYseCkeWeww4EyfvbEgOs46oK/zIBno52/lOzY4/Mj+kYP8xTvl1zadry6s8mUzfqnjQEkGxDFK726LLkrZddPTl+MDSCFqkEqvX48paUJ4Czi/l6PpBCgky52bH9YtQ2H5E7mP/o6197y4iNHBUu8hQHCRsrdXicT28IuzVrX44i6HaYTtsJJUlrAgSGA3HNEikIV+x+FpPKPc0qd5LqHUYUkUz0iJ3xW58Hzct11J95NgNjE8cHyqVz/bGnMM5qZgWIxVaFTaGItaBp1ZbmTT0YwYsXnRgpoaS2VKC7wH1nYr7Usj0OZFFWqToTAzA2D6WH/3IyhtC9AUzBPvWv9pNWKPGWKv0Hlbtu7JLWwBi9ibFdAXxB8gOmetHcepaaNxJrGs21SatmXgoLxkyGGqbLtsmzdPhZ8jXlq3Q6pe2YU/W/d2eq3mDi1EjOPtNk51IFh0ccwO2HiF/", "bm_sv": "B6B1AA3299A60965C948ED678B8F08A1~YAAQh7csMYWnO5qOAQAA/efd2xePLl0i5jEZeJtxgIMdwfuh4DS24Awox59B1+orUi6p2VxmrIT7ShbkpZ7SwkGFfKrC6Dd4SG3dMUNu2vITgH0ua1YzzCBjlx1gbze+I904R92xxTpsrRCOwHpdeiTZ/xcDDUfGdyWOHMUrdVIuzlO/QgvvqZnHpJH8M47fC/8ubrojutfjmbILMOD3LzwEmUZN6+YEzcZYVnOR3v4A5bUB6UMeqhVbODeP~1"}
    dodnow_headers = {"Cache-Control": "no-cache", "Sec-Ch-Ua": "\"Chromium\";v=\"123\", \"Not:A-Brand\";v=\"8\"", "Pragma": "no-cache", "Userbrowserid": "A3dwq8GCGua7sSE6Uie6tQ==", "Sec-Ch-Ua-Mobile": "?0",
                      "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0.6312.88 Safari/537.36", "Authtoken": "ECxw3LAhH08ZoCOhOhduLg==", "Accept": "application/json, text/plain, */*", "If-Modified-Since": "Mon, 26 Jul 1997 05:00:00 GMT", "Sec-Ch-Ua-Platform": "\"Windows\"", "Sec-Fetch-Site": "same-origin", "Sec-Fetch-Mode": "cors", "Sec-Fetch-Dest": "empty", "Referer": "https://a810-dobnow.nyc.gov/publish/Index.html", "Accept-Encoding": "gzip, deflate, br", "Accept-Language": "en-GB,en-US;q=0.9,en;q=0.8", "Priority": "u=1, i"}
    dobnow_res = get(dobnow_url, cookies=dobnow_cookies,
                     headers=dodnow_headers).json()
    PropertyDetails = dobnow_res["PropertyDetails"]
    return PropertyDetails


def getGrade(bin,  bbl):

    url = f"https://a810-dobnow.nyc.gov:443/Publish/WrapperPP/PublicPortal.svc/GetEnergyStarComplience/{bbl}/{bin}"
    cookies = {"_gid": "GA1.2.689702731.1712843572", "RT": "\"z=1&dm=a810-dobnow.nyc.gov&si=e577c804-bfa4-4a11-b664-c04d7b66cc3b&ss=luxqrlny&sl=0&tt=0\"", "_abck": "7B9BD723430B39615EB182E1B879774C~0~YAAQh7csMWKnO5qOAQAA98bd2wsB7Mv+orUtPL+VrMepx0YFXGZNTdn178z/9vkC4ax8LVHMX/c6hIdTzjCrcAdTA2PKak0+juw9zjvSj+vqWHxy/skYJ77V5ldfi2gMBBiwJJwgNIdqOy51hDzK02XZJvaVHrgLo825kQ/WFsw15uYN5OInYAlMYZUTNEQQPVGGzchAn83W8AWnNjCKgR9RbGx0TFkQH/fGxTtp2JJXFII+0PBJCwuqR1lm64/aN/pfShqGFEndCdkiFtTXTguMRzBbObeT1wvSZyQx8r1DPB7J8n0FJObHbvlDivl4WlwuH86RTncfvOvExhoiIHBNsMciQZIGglkhisczFidgiJ1hspigWmXQdP2j+Cigo3GTY7QzXCApKHmcX/VXVYf9Or0I~-1~-1~-1", "bm_sz": "75526566572F8FF567D15EB783B3E3B6~YAAQh7csMWWnO5qOAQAA98bd2xc4eAtLYKEtiXLrJ33xktSNiZVngDELNqt2TtWNC5FLwYoJTgrHATX9j/D8c3ZS/8kzOgCSB6pPMrw2ERmmXzZ1h5PqkRsUoaB0TQ1jNmPrfIsadSs9q6kkfs0GcFyLBJFSAd4L8+pLZKeI8ZA/NieKLdR8DMGJzWF6urW5tydwFbmzNfWqTr+GvXkXsq7zyoyK19AgS7MJXyJeStawDiEPRi0cfzQ13mIhLK/Iy3zfi648JeecDZmj8PSjZ0k9f+WRt+cNfBgyeWxYP5QLWvY22sbSg73CD8CosgLa2TAMN8EsAdOFU0dVZz8XNxG8Sb3vC6O0a1GRe482mKGUVVVIIMfec0wL41JtoTsW5782TInyx8U=~3354949~4408372", "PUNKDDUF": "02ab9cfba6-09e1-43R2b4YO3D2KCbSNHCsQKIlCIHJrRbs-zcnFSIvr7oXtxX0CcG5JUfnnpVmyNv1h3Za28",
               "bm_mi": "E903198E825FBF95CC92E2E311D083C2~YAAQh7csMW6nO5qOAQAAAdLd2xdxOvzkMWLagcXkytlk5QixShzaMncv8zECPUex7QE3+vr6G0F337d3ADRf3Bwx/tP7wEy2c8TC8Vwj8Tv6uHm5ZZcTI2PVrbHGQp84MQ3ZZ7bubifCJ8HszKf9irariI0E8o3DxIgFWpkl3SSiMxohIx4m/MHelpTAV/udgYJAGYSpu11FQgaM3xJ2aZBL7ynWIhqcUNalpC/Mm/fqeB9gKwBIPczV5Hy0xBcA4Pn3OhRvsUEM23Z6/JYYWRStiRpkZThEwgfV5V7P2fZXre2jywSrQpBtvUgVMBw4Oipip56gLjz44cai3KQ6rA==~1", "ASP.NET_SessionId": "aurylxnfhi1u4f102241bqfr", "_ga_863DM8YSJL": "GS1.1.1713085731.9.0.1713085731.0.0.0", "_ga": "GA1.2.1937995955.1712671894", "ak_bmsc": "74E1441C2A1B8732DDBE89DF4BD640D7~000000000000000000000000000000~YAAQh7csMYGnO5qOAQAA19Pd2xd7jT7X9Ocw8RbIsIfgTWkRQYseCkeWeww4EyfvbEgOs46oK/zIBno52/lOzY4/Mj+kYP8xTvl1zadry6s8mUzfqnjQEkGxDFK726LLkrZddPTl+MDSCFqkEqvX48paUJ4Czi/l6PpBCgky52bH9YtQ2H5E7mP/o6197y4iNHBUu8hQHCRsrdXicT28IuzVrX44i6HaYTtsJJUlrAgSGA3HNEikIV+x+FpPKPc0qd5LqHUYUkUz0iJ3xW58Hzct11J95NgNjE8cHyqVz/bGnMM5qZgWIxVaFTaGItaBp1ZbmTT0YwYsXnRgpoaS2VKC7wH1nYr7Usj0OZFFWqToTAzA2D6WH/3IyhtC9AUzBPvWv9pNWKPGWKv0Hlbtu7JLWwBi9ibFdAXxB8gOmetHcepaaNxJrGs21SatmXgoLxkyGGqbLtsmzdPhZ8jXlq3Q6pe2YU/W/d2eq3mDi1EjOPtNk51IFh0ccwO2HiF/", "bm_sv": "B6B1AA3299A60965C948ED678B8F08A1~YAAQh7csMZWnO5qOAQAAuBTe2xd7WsZ0FetqAXRe5+n7f/S+OiS95iTEkdI7LcAODFkwlCnNz4FxRfncOhzu4dXzWfbVHmAuDbtpC89KuNtO0+uk/lQgBR4aLlHUB3oVcAhgzc8feR6AuAa7qcU+9qQ57YBjI1kAiD3T4c9PFI5px+Mvhp0T9/Qyfkz71FUnE9CysDvTMAAsipVxvoWgb6qFdXzKCN3+dPodBIbXauzCdCCT9Heng4H2ePx5~1"}
    headers = {"Cache-Control": "no-cache", "Sec-Ch-Ua": "\"Chromium\";v=\"123\", \"Not:A-Brand\";v=\"8\"", "Pragma": "no-cache", "Userbrowserid": "A3dwq8GCGua7sSE6Uie6tQ==", "Sec-Ch-Ua-Mobile": "?0",
               "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0.6312.88 Safari/537.36", "Authtoken": "ECxw3LAhH08ZoCOhOhduLg==", "Accept": "application/json, text/plain, */*", "If-Modified-Since": "Mon, 26 Jul 1997 05:00:00 GMT", "Sec-Ch-Ua-Platform": "\"Windows\"", "Sec-Fetch-Site": "same-origin", "Sec-Fetch-Mode": "cors", "Sec-Fetch-Dest": "empty", "Referer": "https://a810-dobnow.nyc.gov/publish/Index.html", "Accept-Encoding": "gzip, deflate, br", "Accept-Language": "en-GB,en-US;q=0.9,en;q=0.8", "Priority": "u=1, i"}

    res = get(url, headers=headers, cookies=cookies).json()

    try:
        return res["EnergyStar"]["Gradescore"]
    except:

        return ""
    # print(res)


def takeImage(string, dir, fileName):

    imge_bytes = b64decode(string)
    image_stream = io.BytesIO(imge_bytes)
    img = Image.open(image_stream)

    path = Path(dir)
    path.mkdir(parents=True, exist_ok=True)

    path = f'./{dir}/{fileName}.png'
    img.save(path)

    return path


def estpenalty(html):

    soup = BeautifulSoup(html, 'html.parser')
    select_all_div = soup.findAll('g')

    dic = {}

    for num in range(0, len(select_all_div) - 2):
        text = select_all_div[num].findAll('text')
        dic[text[0].text] = text[-1].text

    return dic


def selenium_chrome(id, value):

    driver = webdriver.Chrome(service=ChromeService(
        ChromeDriverManager().install()))
    driver.get("https://www.be-exchange.org/calculator/")

    sleep(10)

    cal_page = driver.find_element(
        By.CSS_SELECTOR, ".sc-gswNZR.sc-hLBbgP.fkJIIU.hFTahi")
    cal_page.click()

    sleep(5)

    input_tag = driver.find_element(
        By.CSS_SELECTOR, ".MuiInputBase-input.MuiOutlinedInput-input.css-1x5jdmq")
    input_tag.send_keys(id)

    sleep(25)

    select_table = driver.find_element(
        By.CSS_SELECTOR, ".MuiTable-stickyHeader")
    select_table_all_tr = select_table.find_elements(
        By.CSS_SELECTOR, ".MuiTableRow-root.css-18rv9fi")

    if len(select_table_all_tr) == 1:
        driver.close()
        return False

    for tr in select_table_all_tr:

        _data = str(tr.get_attribute("innerText")).lower()

        if value.lower() in _data or id in _data:

            button = tr.find_element(By.TAG_NAME, "button")

            button.click()

            select_building_summary = driver.find_element(
                By.CSS_SELECTOR, ".sc-jfvxQR.ljYkBH")

            inner_Summary_div = select_building_summary.find_elements(
                By.TAG_NAME, 'div')

            dic = {}

            for div in inner_Summary_div:

                select_div = str(div.get_attribute("innerText")).split(':')
                dic[select_div[0]] = select_div[-1]

            Bbl = dic["NYC BBL"].strip().replace("-", "")

            Bin = dic["NYC BIN"].strip().replace("-", "")

            grade = getGrade(Bin, Bbl)

            dic["grade"] = grade

            next_button = driver.find_element(By.CSS_SELECTOR, '.sc-iveFHk')

            next_button.click()

            sleep(5)

            table_data = driver.find_element(By.CSS_SELECTOR, ".table-g")

            html = table_data.get_attribute("innerHTML")
            # est_dic =   estpenalty(html)

            res = estpenalty(html)

            carbon = driver.find_element(
                By.CSS_SELECTOR, '.sc-kgTSHT.cmoktC').screenshot_as_base64

            crbon_path = takeImage(
                string=carbon, dir=value.lower(), fileName="carbon")

            sleep(1)

            cost_page = driver.find_element(By.CSS_SELECTOR, ".fClbry")
            cost_page.click()

            sleep(1)

            cost = driver.find_element(
                By.CSS_SELECTOR, '.sc-kgTSHT.cmoktC').screenshot_as_base64

            # print(carbon, cost)
            cosnt_path = takeImage(
                string=cost, dir=value.lower(), fileName="cost")

            return {

                "dic": dic,
                "res": res,
                "crbon_path": crbon_path,
                "cosnt_path": cosnt_path
            }


def dnbhoovers_search(name):

    burp0_url = "https://app.dnbhoovers.com:443/api/search/typeAheadQuery"
    burp0_cookies = {"apt.uid": "AP-9APX3RS6VUQK-2-1713015149839-81689821.0.2.60d22e54-36fb-4b36-8774-79ca0dd23452", "OSMACH": "95", "OSPWD": "\"7GMElLPOL4bwNWF1xsN/OouhFNU=\"", "LOGONID": "hsetty@zerocircle.eco", "SITE": "174118818", "USERID": "AV_t7tQp5H7LbIBvpfH", "ext_id": "KKVV2AZ4XZFISMVPVI44A1WY3D3JJFER", "JSESSIONID": "E91506CE62DA26297330988366590267.prd1-l-app02", "ak_bmsc": "D9BE76DB8A8FBBD984A0C6F2DF8753C0~000000000000000000000000000000~YAAQ7ulUuN1J/8uOAQAACQOO2xcDnA4NLgZssNeY56ms/eUgyW+CfsrdgjoIHXZnPm9TcZBfGoZ3fcIVweXgQrW7q3jsPvpPdTSvp67EoCYhwTJ7LT1uWr+aDi+b1F7ofAiJoe12GdM7jK5G+I23KB85TSI90JlWRzLc5yZSkdzvgzFNtqzl7p9/iZJemh6r14RWDeJSJn/AwOhCIcjk61VuRaxUaCd2GNpdKDQJx/vC8PSqntUmWZyEYRKkKiuuJ076nfGS+sIXcgb8rwxeiKnFOFhL2iDCjXRwmyobVfsubHJbtTTWjVR6M6rhFhcgN162TT7EIhIkRPmskk/MS0BbQuXmnd6nBoGNrRrYyAQSpjLmf39YkMAgrgjuNl4A+6oKIyhRWqKOzSlBsA==",
                     "bm_sv": "3B20A6190F55FC5E284F4D9254404B14~YAAQ7ulUuJmkAcyOAQAADQTl2xcI1i5BiRitY0MNQy4Jdq8c13H7S9pdm5n1eU5U5ZOr3/fORpSZOVxQStgFYXvC7I/nuR2f3BKiD17CVnO7gPklUjkPGPf1mdpfAi0lSu4a1EEFWke75DuRDMU/Hl0M47woO1w/hRGfye/UsqnNCm78uQAUOv4gRbpJzpsUTmlea3NOkgxMqPilhK+rgU3YbqTl0gCiMzGc2fAqoHriUWu0Jnu+h8p6ZHtN58qbogIZzEo=~1", "GCILB": "\"475aa7e2cb5a6bac\"", "apt.sid": "AP-9APX3RS6VUQK-2-1713086395828-86933697"}
    burp0_headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:124.0) Gecko/20100101 Firefox/124.0", "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8", "Accept-Language": "en-US,en;q=0.5", "Accept-Encoding": "gzip, deflate, br",
                     "Referer": "https://app.dnbhoovers.com/company/a72f0578-f414-3042-ae6c-9c988c020c70", "Upgrade-Insecure-Requests": "1", "Sec-Fetch-Dest": "document", "Sec-Fetch-Mode": "navigate", "Sec-Fetch-Site": "same-origin", "Sec-Fetch-User": "?1", "Te": "trailers"}
    burp0_json = {"primitives": ["company"], "query": name}
    data = post(burp0_url, headers=burp0_headers,
                cookies=burp0_cookies, json=burp0_json).json()
    list_ = data['company']['searchResults']['results']

    if len(list_) > 0:
        return list_[0]['id']

    return False


def get_company(id):
    url = "https://app.dnbhoovers.com:443/company/" + str(id)
    cookis = {"apt.uid": "AP-9APX3RS6VUQK-2-1713015149839-81689821.0.2.60d22e54-36fb-4b36-8774-79ca0dd23452", "OSMACH": "95", "OSPWD": "\"7GMElLPOL4bwNWF1xsN/OouhFNU=\"", "LOGONID": "hsetty@zerocircle.eco", "SITE": "174118818", "USERID": "AV_t7tQp5H7LbIBvpfH", "ext_id": "KKVV2AZ4XZFISMVPVI44A1WY3D3JJFER", "JSESSIONID": "E91506CE62DA26297330988366590267.prd1-l-app02", "ak_bmsc": "D9BE76DB8A8FBBD984A0C6F2DF8753C0~000000000000000000000000000000~YAAQ7ulUuN1J/8uOAQAACQOO2xcDnA4NLgZssNeY56ms/eUgyW+CfsrdgjoIHXZnPm9TcZBfGoZ3fcIVweXgQrW7q3jsPvpPdTSvp67EoCYhwTJ7LT1uWr+aDi+b1F7ofAiJoe12GdM7jK5G+I23KB85TSI90JlWRzLc5yZSkdzvgzFNtqzl7p9/iZJemh6r14RWDeJSJn/AwOhCIcjk61VuRaxUaCd2GNpdKDQJx/vC8PSqntUmWZyEYRKkKiuuJ076nfGS+sIXcgb8rwxeiKnFOFhL2iDCjXRwmyobVfsubHJbtTTWjVR6M6rhFhcgN162TT7EIhIkRPmskk/MS0BbQuXmnd6nBoGNrRrYyAQSpjLmf39YkMAgrgjuNl4A+6oKIyhRWqKOzSlBsA==",
              "bm_sv": "3B20A6190F55FC5E284F4D9254404B14~YAAQ7ulUuJmkAcyOAQAADQTl2xcI1i5BiRitY0MNQy4Jdq8c13H7S9pdm5n1eU5U5ZOr3/fORpSZOVxQStgFYXvC7I/nuR2f3BKiD17CVnO7gPklUjkPGPf1mdpfAi0lSu4a1EEFWke75DuRDMU/Hl0M47woO1w/hRGfye/UsqnNCm78uQAUOv4gRbpJzpsUTmlea3NOkgxMqPilhK+rgU3YbqTl0gCiMzGc2fAqoHriUWu0Jnu+h8p6ZHtN58qbogIZzEo=~1", "GCILB": "\"475aa7e2cb5a6bac\"", "apt.sid": "AP-9APX3RS6VUQK-2-1713086395828-86933697"}
    headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:124.0) Gecko/20100101 Firefox/124.0", "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8", "Accept-Language": "en-US,en;q=0.5", "Accept-Encoding": "gzip, deflate, br",
               "Referer": "https://app.dnbhoovers.com/company/a72f0578-f414-3042-ae6c-9c988c020c70", "Upgrade-Insecure-Requests": "1", "Sec-Fetch-Dest": "document", "Sec-Fetch-Mode": "navigate", "Sec-Fetch-Site": "same-origin", "Sec-Fetch-User": "?1", "Te": "trailers"}
    res = get(url, headers=headers, cookies=cookis).text

    soup = BeautifulSoup(res, 'html.parser').find(
        'script', {'id': "PrimitiveJson"}
    )

    ctx = compile(soup.text)
    initialLoad = ctx.eval("initialLoad")

    selectd_data = initialLoad["primitiveDetails"]['primitive']

    data = []
    company = selectd_data["companyName"]
    data.append(["compnay name", company])
    phone = selectd_data["phone"]
    data.append(["phone", phone])

    address = selectd_data["addresses"][0]
    res_address = address["address1"] + ", " + \
        address["city"] + ", " + address["postalCode"]
    data.append(["addrss", res_address, company])
    ultimateParentDunsNumber = selectd_data["ultimateParentDunsNumber"]
    data.append(["ultimateParentDunsNumber", ultimateParentDunsNumber])
    latitude = address["latitude"]
    data.append(["latitude", latitude])
    longitude = address["longitude"]
    data.append(["longitude", longitude])
    industries = selectd_data["industries"]
    data.append(["industries", industries])
    yearFounded = selectd_data["yearFounded"]
    data.append(["yearFounded", yearFounded])

    for _data in industries:

        if _data["industry"]:

            data.append([_data["industry"]["source"], _data["industry"]["code"] +
                         _data["industry"]["shortDescription"]])

    county = address["county"]["name"]
    data.append(["county", county])
    ultimateParentDunsNumber = selectd_data["ultimateParentDunsNumber"]
    data.append(["ultimateParentDunsNumber", ultimateParentDunsNumber])
    domesticUltimateParent = selectd_data["domesticUltimateParent"]['dunsNumber']
    data.append(["domesticUltimateParent", domesticUltimateParent])

    return data


def getCatact(id, name):

    burp0_url = "https://app.dnbhoovers.com:443/api/search"
    burp0_cookies = {"apt.uid": "AP-9APX3RS6VUQK-2-1713015149839-81689821.0.2.60d22e54-36fb-4b36-8774-79ca0dd23452", "OSMACH": "95", "OSPWD": "\"7GMElLPOL4bwNWF1xsN/OouhFNU=\"", "LOGONID": "hsetty@zerocircle.eco", "SITE": "174118818", "USERID": "AV_t7tQp5H7LbIBvpfH", "ext_id": "KKVV2AZ4XZFISMVPVI44A1WY3D3JJFER", "JSESSIONID": "E91506CE62DA26297330988366590267.prd1-l-app02", "ak_bmsc": "D9BE76DB8A8FBBD984A0C6F2DF8753C0~000000000000000000000000000000~YAAQ7ulUuN1J/8uOAQAACQOO2xcDnA4NLgZssNeY56ms/eUgyW+CfsrdgjoIHXZnPm9TcZBfGoZ3fcIVweXgQrW7q3jsPvpPdTSvp67EoCYhwTJ7LT1uWr+aDi+b1F7ofAiJoe12GdM7jK5G+I23KB85TSI90JlWRzLc5yZSkdzvgzFNtqzl7p9/iZJemh6r14RWDeJSJn/AwOhCIcjk61VuRaxUaCd2GNpdKDQJx/vC8PSqntUmWZyEYRKkKiuuJ076nfGS+sIXcgb8rwxeiKnFOFhL2iDCjXRwmyobVfsubHJbtTTWjVR6M6rhFhcgN162TT7EIhIkRPmskk/MS0BbQuXmnd6nBoGNrRrYyAQSpjLmf39YkMAgrgjuNl4A+6oKIyhRWqKOzSlBsA==",
                     "bm_sv": "3B20A6190F55FC5E284F4D9254404B14~YAAQ7ulUuJmkAcyOAQAADQTl2xcI1i5BiRitY0MNQy4Jdq8c13H7S9pdm5n1eU5U5ZOr3/fORpSZOVxQStgFYXvC7I/nuR2f3BKiD17CVnO7gPklUjkPGPf1mdpfAi0lSu4a1EEFWke75DuRDMU/Hl0M47woO1w/hRGfye/UsqnNCm78uQAUOv4gRbpJzpsUTmlea3NOkgxMqPilhK+rgU3YbqTl0gCiMzGc2fAqoHriUWu0Jnu+h8p6ZHtN58qbogIZzEo=~1", "GCILB": "\"475aa7e2cb5a6bac\"", "apt.sid": "AP-9APX3RS6VUQK-2-1713086395828-86933697"}
    burp0_headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:124.0) Gecko/20100101 Firefox/124.0", "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8", "Accept-Language": "en-US,en;q=0.5", "Accept-Encoding": "gzip, deflate, br",
                     "Referer": "https://app.dnbhoovers.com/company/a72f0578-f414-3042-ae6c-9c988c020c70", "Upgrade-Insecure-Requests": "1", "Sec-Fetch-Dest": "document", "Sec-Fetch-Mode": "navigate", "Sec-Fetch-Site": "same-origin", "Sec-Fetch-User": "?1", "Te": "trailers"}
    burp0_json = {"aggs": {"contactCompanyCount": {}}, "filters": [{"contactCompanyIdFacet": {"exclude": False, "label": "Contact Companies", "type": "", "valueLabels": [name], "values": [
        id]}}], "from": 0, "query": "", "searchWeight": 0, "size": 25, "sortBy": [{"contact": [{"CONTACT_LEVEL": "asc"}]}], "types": ["contact"], "valueLabels": {id: name}, "version": 2}
    res = post(burp0_url, headers=burp0_headers,
               cookies=burp0_cookies, json=burp0_json).json()

    selectd_data = res["searchResults"]["results"]

    final_res = []

    for data in selectd_data:

        final_res.append(
            ["", "fullName", data["firstName"] + " " + data["lastName"]])

        final_res.append(["", "job title", data["title"]])

        final_res.append(
            ["", "linked url", data["linkedInPublicProfileUrl"]]
        )
        final_res.append([
            "", "phone", data["company"]["phone"]
        ])

    return final_res


def dnbhoovers(owner):
    id = dnbhoovers_search(owner)

    print(id)

    if id:
        company = get_company(id)
        contact = getCatact(id, owner)

        return company + contact

    # if id:
    #     get_company(id)


for num in range(2, len(ld["A"])):

    iter_wb = Workbook()
    iter_active = iter_wb.active
    value = [data.value for data in ld[num:num]]

    property_address = value[1]
    door_num = ""
    street_address = ""

    owner_name = dnbhoovers(str(value[10]))

    for str_ in property_address.split(" "):

        if str(str_).isnumeric() or str(str_).split("-")[0].isnumeric():
            door_num = str_.split("-")[0]
        else:
            street_address += str_ + " "

    suffix_url = "1|" + door_num + "|" + street_address.strip() + "|1"
    url_encode_data = suffix_url.replace("|", "%7C").replace(" ", "%20")

    donow_res = getDobnow(url_encode_data)
    Bin = donow_res["BIN"]

    count = 0
    found = True

    data = ""

    while count < 5 and found:

        if count % 2:

            data = selenium_chrome(Bin, value[1])
            if data:
                break
        else:
            data = selenium_chrome(property_address, value[1])

            if data:
                break
        count += 1

    iter_active.append(["dobnow"])
    for data in donow_res:
        iter_active.append([data, donow_res[data]])

    iter_active.append(["Be Exchange calculator"])

    if 'str' not in type(data):

        for data in data["dic"]:

            print([data, data["dic"]['data']])

            iter_active.append([data, data["dic"]['data']])

        crbon = drawing.Image(data["crbon_path"])
        crbon.anchor(iter_active.cell("M2"))
        iter_active.add_image(crbon)

        cosnt = drawing.Image(data["crbon_path"])
        cosnt.anchor(iter_active.cell("M102"))
        iter_active.add_image(cosnt)

    iter_active.append(["dnbhoovers"])

    for _ in owner_name:
        if "industries" in _:
            for owner_name in _[1]:
                if _["industry"]:

                    iter_active.append([_["industry"]["source"], _[
                                       "industry"]["shortDescription"]])
                else:
                    iter_active.append(_)

    iter_wb.save(filename=property_address)
